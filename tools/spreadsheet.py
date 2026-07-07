"""Bounded reading, viewing, and creation of CSV and Excel spreadsheets."""

from __future__ import annotations

import csv
import json
import math
import os
import re
import tempfile
from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Callable


SUPPORTED_EXTENSIONS = {".csv", ".xls", ".xlsx"}
MAX_FILE_BYTES = 25 * 1024 * 1024
MAX_SHEETS = 20
MAX_CREATE_ROWS = 10_000
MAX_CREATE_COLUMNS = 256
MAX_CREATE_CELLS = 100_000
MAX_READ_ROWS = 200
MAX_READ_COLUMNS = 100
MAX_SCAN_CELLS = 200_000
MAX_QUERY_MATCHES = 50
MAX_CSV_CELLS = 200_000
CSV_DELIMITERS = {",", ";", "\t", "|"}
_CELL_RANGE_RE = re.compile(r"^([A-Za-z]+)([1-9]\d*)(?::([A-Za-z]+)([1-9]\d*))?$")
_INVALID_SHEET_CHARS = re.compile(r"[\\/*?:\[\]]")


@dataclass
class _Sheet:
    name: str
    row_count: int
    column_count: int
    value_at: Callable[[int, int], Any]


def _json(value: dict) -> str:
    return json.dumps(value, ensure_ascii=False, allow_nan=False)


def _bounded_int(value: Any, default: int, minimum: int, maximum: int) -> int:
    try:
        parsed = int(value) if value is not None else default
    except (TypeError, ValueError):
        parsed = default
    return max(minimum, min(maximum, parsed))


def _serialize_cell(value: Any) -> Any:
    if value is None or isinstance(value, (str, bool, int)):
        return value
    if isinstance(value, float):
        return value if math.isfinite(value) else str(value)
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    return str(value)


def _column_number(letters: str) -> int:
    result = 0
    for character in letters.upper():
        if not "A" <= character <= "Z":
            raise ValueError(f"Invalid column: {letters}")
        result = result * 26 + ord(character) - ord("A") + 1
    return result


def _column_letters(number: int) -> str:
    value = number
    result = []
    while value:
        value, remainder = divmod(value - 1, 26)
        result.append(chr(ord("A") + remainder))
    return "".join(reversed(result)) or "A"


def _parse_range(cell_range: str | None, sheet: _Sheet) -> tuple[int, int, int, int]:
    if not cell_range:
        return 1, 1, sheet.row_count, sheet.column_count
    match = _CELL_RANGE_RE.fullmatch(str(cell_range).strip())
    if not match:
        raise ValueError("cell_range must use A1 or A1:D20 notation")
    start_column = _column_number(match.group(1))
    start_row = int(match.group(2))
    end_column = _column_number(match.group(3) or match.group(1))
    end_row = int(match.group(4) or match.group(2))
    if start_row > end_row:
        start_row, end_row = end_row, start_row
    if start_column > end_column:
        start_column, end_column = end_column, start_column
    return start_row, start_column, end_row, end_column


def _trim_matrix(rows: list[list[Any]]) -> list[list[Any]]:
    trimmed = []
    for row in rows:
        values = list(row)
        while values and values[-1] is None:
            values.pop()
        trimmed.append(values)
    while trimmed and not trimmed[-1]:
        trimmed.pop()
    return trimmed


def _matrix(
    sheet: _Sheet,
    bounds: tuple[int, int, int, int],
    max_rows: int,
    max_columns: int,
) -> tuple[list[list[Any]], dict]:
    start_row, start_column, requested_end_row, requested_end_column = bounds
    actual_end_row = min(requested_end_row, sheet.row_count, start_row + max_rows - 1)
    actual_end_column = min(requested_end_column, sheet.column_count, start_column + max_columns - 1)
    if start_row > sheet.row_count or start_column > sheet.column_count:
        return [], {
            "range": f"{_column_letters(start_column)}{start_row}",
            "truncated": False,
        }
    rows = [
        [_serialize_cell(sheet.value_at(row, column)) for column in range(start_column, actual_end_column + 1)]
        for row in range(start_row, actual_end_row + 1)
    ]
    truncated = actual_end_row < min(requested_end_row, sheet.row_count) or actual_end_column < min(requested_end_column, sheet.column_count)
    return _trim_matrix(rows), {
        "range": f"{_column_letters(start_column)}{start_row}:{_column_letters(actual_end_column)}{actual_end_row}",
        "truncated": truncated,
    }


def _open_xlsx(file_path: Path, data_only: bool) -> tuple[list[_Sheet], Callable[[], None]]:
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("Missing spreadsheet dependency. Run: pip install openpyxl") from exc
    workbook = openpyxl.load_workbook(
        file_path,
        read_only=True,
        data_only=data_only,
        keep_links=False,
    )
    sheets = []
    for worksheet in workbook.worksheets[:MAX_SHEETS]:
        sheets.append(_Sheet(
            name=worksheet.title,
            row_count=int(worksheet.max_row or 0),
            column_count=int(worksheet.max_column or 0),
            value_at=lambda row, column, ws=worksheet: ws.cell(row=row, column=column).value,
        ))
    return sheets, workbook.close


def _open_xls(file_path: Path, data_only: bool) -> tuple[list[_Sheet], Callable[[], None]]:
    del data_only  # Legacy .xls readers expose stored values rather than a data-only mode.
    try:
        import xlrd
    except ImportError as exc:
        raise RuntimeError("Missing spreadsheet dependency. Run: pip install xlrd") from exc
    workbook = xlrd.open_workbook(file_path, on_demand=True)

    def cell_value(worksheet, row: int, column: int):
        cell = worksheet.cell(row - 1, column - 1)
        if cell.ctype == xlrd.XL_CELL_DATE:
            return xlrd.xldate_as_datetime(cell.value, workbook.datemode)
        if cell.ctype in {xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK}:
            return None
        if cell.ctype == xlrd.XL_CELL_BOOLEAN:
            return bool(cell.value)
        return cell.value

    sheets = []
    for worksheet in workbook.sheets()[:MAX_SHEETS]:
        sheets.append(_Sheet(
            name=worksheet.name,
            row_count=worksheet.nrows,
            column_count=worksheet.ncols,
            value_at=lambda row, column, ws=worksheet: cell_value(ws, row, column),
        ))
    return sheets, workbook.release_resources


def _validate_delimiter(delimiter: str | None) -> str | None:
    if delimiter is None or delimiter == "":
        return None
    value = str(delimiter)
    if value == r"\t":
        value = "\t"
    if value not in CSV_DELIMITERS:
        raise ValueError("delimiter must be one of comma, semicolon, tab, or pipe")
    return value


def _open_csv(file_path: Path, data_only: bool, delimiter: str | None) -> tuple[list[_Sheet], Callable[[], None]]:
    del data_only
    selected_delimiter = _validate_delimiter(delimiter)
    with open(file_path, "r", encoding="utf-8-sig", newline="") as stream:
        if selected_delimiter is None:
            sample = stream.read(8192)
            stream.seek(0)
            try:
                selected_delimiter = csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
            except csv.Error:
                selected_delimiter = ","
        reader = csv.reader(stream, delimiter=selected_delimiter)
        rows = []
        total_cells = 0
        column_count = 0
        for row in reader:
            total_cells += len(row)
            if total_cells > MAX_CSV_CELLS:
                raise ValueError(f"CSV exceeds the {MAX_CSV_CELLS}-cell read limit")
            column_count = max(column_count, len(row))
            rows.append(row)

    def value_at(row: int, column: int):
        values = rows[row - 1]
        return values[column - 1] if column <= len(values) else None

    return [_Sheet(
        name="CSV",
        row_count=len(rows),
        column_count=column_count,
        value_at=value_at,
    )], lambda: None


def _open_workbook(
    file_path: Path,
    data_only: bool,
    delimiter: str | None = None,
) -> tuple[list[_Sheet], Callable[[], None]]:
    extension = file_path.suffix.lower()
    if extension == ".xlsx":
        return _open_xlsx(file_path, data_only)
    if extension == ".xls":
        return _open_xls(file_path, data_only)
    return _open_csv(file_path, data_only, delimiter)


def _select_sheet(sheets: list[_Sheet], name: str | None) -> _Sheet:
    if not sheets:
        raise ValueError("Workbook contains no worksheets")
    if not name:
        return sheets[0]
    for sheet in sheets:
        if sheet.name.casefold() == str(name).strip().casefold():
            return sheet
    raise ValueError(f"Worksheet not found: {name}")


def _validate_input_sheets(sheets: Any, extension: str) -> list[tuple[str, list[list[Any]]]]:
    if not isinstance(sheets, list) or not sheets:
        raise ValueError("sheets must be a non-empty array")
    if len(sheets) > MAX_SHEETS:
        raise ValueError(f"A workbook may contain at most {MAX_SHEETS} worksheets")
    validated = []
    names = set()
    total_cells = 0
    format_row_limit = 65_536 if extension == ".xls" else MAX_CREATE_ROWS
    for index, item in enumerate(sheets):
        if not isinstance(item, dict):
            raise ValueError(f"sheets[{index}] must be an object")
        name = str(item.get("name") or f"Sheet{index + 1}").strip()
        if not name or len(name) > 31 or _INVALID_SHEET_CHARS.search(name):
            raise ValueError(f"Invalid worksheet name: {name!r}")
        if name.casefold() in names:
            raise ValueError(f"Duplicate worksheet name: {name}")
        names.add(name.casefold())
        rows = item.get("rows", [])
        if not isinstance(rows, list):
            raise ValueError(f"rows for worksheet '{name}' must be an array")
        if len(rows) > min(MAX_CREATE_ROWS, format_row_limit):
            raise ValueError(f"Worksheet '{name}' exceeds the {min(MAX_CREATE_ROWS, format_row_limit)}-row creation limit")
        normalized_rows = []
        for row_index, row in enumerate(rows):
            if not isinstance(row, list):
                raise ValueError(f"Worksheet '{name}' row {row_index + 1} must be an array")
            if len(row) > MAX_CREATE_COLUMNS:
                raise ValueError(f"Worksheet '{name}' row {row_index + 1} exceeds {MAX_CREATE_COLUMNS} columns")
            normalized = []
            for value in row:
                if value is not None and not isinstance(value, (str, bool, int, float)):
                    raise ValueError("Cell values must be strings, numbers, booleans, or null")
                if isinstance(value, float) and not math.isfinite(value):
                    raise ValueError("Cell numbers must be finite")
                normalized.append(value)
            normalized_rows.append(normalized)
            total_cells += len(normalized)
            if total_cells > MAX_CREATE_CELLS:
                raise ValueError(f"Workbook exceeds the {MAX_CREATE_CELLS}-cell creation limit")
        validated.append((name, normalized_rows))
    return validated


def _create_xlsx(file_path: Path, sheets: list[tuple[str, list[list[Any]]]], allow_formulas: bool) -> None:
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("Missing spreadsheet dependency. Run: pip install openpyxl") from exc
    workbook = openpyxl.Workbook(write_only=False)
    for index, (name, rows) in enumerate(sheets):
        worksheet = workbook.active if index == 0 else workbook.create_sheet()
        worksheet.title = name
        for row_index, row in enumerate(rows, start=1):
            for column_index, value in enumerate(row, start=1):
                cell = worksheet.cell(row=row_index, column=column_index, value=value)
                if isinstance(value, str) and value.startswith("=") and not allow_formulas:
                    cell.data_type = "s"
    workbook.save(file_path)
    workbook.close()


def _create_xls(file_path: Path, sheets: list[tuple[str, list[list[Any]]]], allow_formulas: bool) -> None:
    try:
        import xlwt
    except ImportError as exc:
        raise RuntimeError("Missing spreadsheet dependency. Run: pip install xlwt") from exc
    workbook = xlwt.Workbook()
    for name, rows in sheets:
        worksheet = workbook.add_sheet(name)
        for row_index, row in enumerate(rows):
            for column_index, value in enumerate(row):
                if allow_formulas and isinstance(value, str) and value.startswith("="):
                    worksheet.write(row_index, column_index, xlwt.Formula(value[1:]))
                else:
                    worksheet.write(row_index, column_index, "" if value is None else value)
    workbook.save(str(file_path))


def _create_csv(
    file_path: Path,
    sheets: list[tuple[str, list[list[Any]]]],
    allow_formulas: bool,
    delimiter: str | None,
) -> None:
    if len(sheets) != 1:
        raise ValueError("CSV creation accepts exactly one worksheet/rows array")
    selected_delimiter = _validate_delimiter(delimiter) or ","
    with open(file_path, "w", encoding="utf-8", newline="") as stream:
        writer = csv.writer(stream, delimiter=selected_delimiter, lineterminator="\n")
        for row in sheets[0][1]:
            safe_row = []
            for value in row:
                if not allow_formulas and isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
                    value = "'" + value
                safe_row.append("" if value is None else value)
            writer.writerow(safe_row)


def _create_workbook(
    file_path: Path,
    sheets: Any,
    overwrite: bool,
    allow_formulas: bool,
    delimiter: str | None,
) -> dict:
    extension = file_path.suffix.lower()
    validated_sheets = _validate_input_sheets(sheets, extension)
    existed = file_path.exists()
    if existed and not overwrite:
        raise FileExistsError(f"File already exists and overwrite=false: {file_path}")
    if not file_path.parent.is_dir():
        raise ValueError(f"Parent directory does not exist: {file_path.parent}")
    handle, temporary_name = tempfile.mkstemp(prefix=f".{file_path.stem}-", suffix=extension, dir=file_path.parent)
    os.close(handle)
    temporary = Path(temporary_name)
    try:
        if extension == ".xlsx":
            _create_xlsx(temporary, validated_sheets, allow_formulas)
        elif extension == ".xls":
            _create_xls(temporary, validated_sheets, allow_formulas)
        else:
            _create_csv(temporary, validated_sheets, allow_formulas, delimiter)
        os.replace(temporary, file_path)
    finally:
        if temporary.exists():
            temporary.unlink()
    return {
        "ok": True,
        "action": "create",
        "file": str(file_path),
        "format": extension[1:],
        "sheet_names": [name for name, _ in validated_sheets],
        "sheet_count": len(validated_sheets),
        "row_counts": {name: len(rows) for name, rows in validated_sheets},
        "formulas_enabled": allow_formulas,
        "overwritten": existed,
        **({"delimiter": _validate_delimiter(delimiter) or ","} if extension == ".csv" else {}),
    }


def spreadsheet(
    action: str,
    file_path: str,
    sheet: str | None = None,
    cell_range: str | None = None,
    query: str | None = None,
    sheets: list[dict] | None = None,
    rows: list[list[Any]] | None = None,
    max_rows: int | str = 50,
    max_columns: int | str = 30,
    data_only: bool = False,
    overwrite: bool = False,
    allow_formulas: bool = False,
    delimiter: str | None = None,
    confirmed: bool = False,
) -> str:
    """View, read, query, or create a bounded .csv/.xls/.xlsx workbook."""
    action = str(action or "").strip().lower()
    if action not in {"view", "read", "create"}:
        return _json({"error": "action must be view, read, or create"})
    raw_path = str(file_path or "").strip()
    if not raw_path:
        return _json({"error": "file_path is required"})
    path = Path(raw_path).expanduser()
    if not path.is_absolute():
        path = Path.cwd() / path
    path = path.resolve()
    extension = path.suffix.lower()
    if extension not in SUPPORTED_EXTENSIONS:
        return _json({"error": "Only .csv, .xls, and .xlsx files are supported"})

    try:
        if action == "create":
            if confirmed is not True:
                return _json({
                    "error": "Creating a spreadsheet requires explicit user approval.",
                    "required": "Call again with confirmed=true only after the user requests creation.",
                })
            creation_sheets = sheets
            if creation_sheets is None and rows is not None:
                creation_sheets = [{"name": "CSV" if extension == ".csv" else "Sheet1", "rows": rows}]
            return _json(_create_workbook(
                path,
                creation_sheets,
                bool(overwrite),
                bool(allow_formulas),
                delimiter,
            ))

        if not path.is_file():
            return _json({"error": f"Spreadsheet not found: {path}"})
        file_size = path.stat().st_size
        if file_size > MAX_FILE_BYTES:
            return _json({
                "error": f"Spreadsheet exceeds the {MAX_FILE_BYTES}-byte read limit",
                "size_bytes": file_size,
            })
        row_limit = _bounded_int(max_rows, 50, 1, MAX_READ_ROWS)
        column_limit = _bounded_int(max_columns, 30, 1, MAX_READ_COLUMNS)
        workbook_sheets, close = _open_workbook(path, bool(data_only), delimiter)
        try:
            base = {
                "ok": True,
                "action": action,
                "file": str(path),
                "format": extension[1:],
                "size_bytes": file_size,
                "sheet_names": [value.name for value in workbook_sheets],
                "sheet_count": len(workbook_sheets),
                **({"delimiter": _validate_delimiter(delimiter)} if extension == ".csv" and delimiter else {}),
            }
            if action == "view":
                selected = [_select_sheet(workbook_sheets, sheet)] if sheet else workbook_sheets
                previews = []
                for value in selected:
                    rows, info = _matrix(
                        value,
                        (1, 1, value.row_count, value.column_count),
                        min(row_limit, 25),
                        min(column_limit, 20),
                    )
                    previews.append({
                        "name": value.name,
                        "row_count": value.row_count,
                        "column_count": value.column_count,
                        "preview": rows,
                        "preview_range": info["range"],
                        "preview_truncated": info["truncated"],
                    })
                return _json({**base, "sheets": previews})

            selected_sheets = [_select_sheet(workbook_sheets, sheet)] if sheet else workbook_sheets
            if query is not None and str(query).strip():
                needle = str(query).strip().casefold()
                if len(needle) > 1000:
                    raise ValueError("query exceeds the 1000-character limit")
                matches = []
                scanned = 0
                scan_truncated = False
                for value in selected_sheets:
                    bounds = _parse_range(cell_range, value)
                    start_row, start_column, end_row, end_column = bounds
                    for row in range(start_row, min(end_row, value.row_count) + 1):
                        for column in range(start_column, min(end_column, value.column_count) + 1):
                            scanned += 1
                            if scanned > MAX_SCAN_CELLS:
                                scan_truncated = True
                                break
                            cell_value = _serialize_cell(value.value_at(row, column))
                            if cell_value is not None and needle in str(cell_value).casefold():
                                matches.append({
                                    "sheet": value.name,
                                    "cell": f"{_column_letters(column)}{row}",
                                    "value": cell_value,
                                })
                                if len(matches) >= MAX_QUERY_MATCHES:
                                    scan_truncated = True
                                    break
                        if scan_truncated:
                            break
                    if scan_truncated:
                        break
                return _json({
                    **base,
                    "query": str(query).strip(),
                    "matches": matches,
                    "match_count": len(matches),
                    "cells_scanned": min(scanned, MAX_SCAN_CELLS),
                    "truncated": scan_truncated,
                })

            selected = _select_sheet(workbook_sheets, sheet)
            bounds = _parse_range(cell_range, selected)
            rows, info = _matrix(selected, bounds, row_limit, column_limit)
            return _json({
                **base,
                "sheet": selected.name,
                "row_count": selected.row_count,
                "column_count": selected.column_count,
                "range": info["range"],
                "rows": rows,
                "truncated": info["truncated"],
                "data_only": bool(data_only),
            })
        finally:
            close()
    except (OSError, ValueError, RuntimeError, FileExistsError) as exc:
        return _json({"error": str(exc), "action": action, "file": str(path)})
